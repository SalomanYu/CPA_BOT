import xlrd, xlsxwriter
import csv
import os
import openpyxl
import time
import gspread
from oauth2client.service_account import ServiceAccountCredentials

from sys import platform

if platform == 'win32':
    import ctypes
    kernel32 = ctypes.windll.kernel32
    kernel32.SetConsoleMode(kernel32.GetStdHandle(-11), 7)



SUCCESS_MESSAGE = '\033[2;30;42m [SUCCESS] \033[0;0m' 
WARNING_MESSAGE = '\033[2;30;43m [WARNING] \033[0;0m'
ERROR_MESSAGE = '\033[2;30;41m [ ERROR ] \033[0;0m'

class  CPA:
    def __init__(self, google_sheet_id):
        self.google_sheet_id = google_sheet_id
        self.NONE_EXIST_NIKS = []
        self.NONE_EXIST_ARTICLES = []
        os.makedirs('Result', exist_ok=True)
    
    def run(self):
        # 1. Скачиваем содержимое определенной таблицы
        spread = self.auth_spread(self.google_sheet_id)
        self.worksheet = spread.get_worksheet(6)
        self.download_sheet(self.worksheet)

        # 2. Собираем необходимые данные из вводного excel-файла
        try:
            excel_file = self.extract_excel_file()
        except UnboundLocalError:
            print(ERROR_MESSAGE+'\t В папке Upload не найдено ни одного файла с расширением .xlsx')
            return
        self.all_niks, self.all_articles = self.collect_all_niks_and_articles(excel_file)

        self.articles_with_status = self.get_articlles_with_statuses(list(set(self.all_articles)))
        self.find_all_articles_in_sheet(self.articles_with_status)
        self.create_non_existent_FILE(data=self.NONE_EXIST_ARTICLES, filename='NONE ARTICLES')
        self.create_non_existent_TABLE('NONE ARTICLES',  self.nik_col)
        
        #niks
        freq_niks = self.get_freq_dict(self.all_niks)
        self.find_niks_in_sheet(freq_niks)
        self.create_non_existent_FILE(data=self.NONE_EXIST_NIKS, filename='NONE NIKS')
        self.create_non_existent_TABLE('NONE NIKS',  self.nik_col)
                


    def auth_spread(self, table_id):
        scope = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_name('Service Accounts/morbot-338716-b219142d9c70.json', scope)

        gc = gspread.authorize(credentials)
        spread = gc.open_by_key(table_id)
        print(SUCCESS_MESSAGE+'\t Подключились к гугл таблицам')

        return spread


    def download_sheet(self, worksheet):
        filename = 'Result/downloaded_sheet' + '.xlsx'
        workbook = xlsxwriter.Workbook(filename)
        sheet = workbook.add_worksheet()
        all_values = worksheet.get_all_values()
        for row_num,row_data in enumerate(all_values):
            for col_num, col_data in enumerate(row_data):
                sheet.write(row_num, col_num, col_data)
        print(SUCCESS_MESSAGE+'\t Сохранили таблиицу в отдельный файл')
        workbook.close()

    def extract_excel_file(self) -> str:
        upload_folder = 'Upload'
        os.makedirs(upload_folder, exist_ok=True)
        for file in os.listdir(upload_folder):
            full_path = os.path.join(upload_folder, file)
            if file.endswith('.xlsx'):
                excel_file = full_path
            elif file.endswith('.xls'):
                new_filename = os.path.join(upload_folder, file.split('.')[0] + '.xlsx')
                os.rename(full_path, new_filename)
                excel_file = os.path.join(upload_folder, new_filename)
                print(WARNING_MESSAGE+'\t Поменяли расширение excel файла')
        print(SUCCESS_MESSAGE+'\t Нашли таблицу с исходными значениями из папки ',  upload_folder)
        return excel_file 

    def collect_all_niks_and_articles(self, excel_file:str) -> tuple:
        book_reader = xlrd.open_workbook(excel_file)
        self.sheet_reader = book_reader.sheet_by_index(0)
        self.table_titles = self.sheet_reader.row_values(0)
        
        all_articles = None # Делаем это, чтобы не перезаписывать переменную, т.к существует две колонки "Товары" 

        for title_num in range(len(self.table_titles)):
            if self.table_titles[title_num] == 'nik товара':
                # print(title_num)
                self.nik_col = title_num
            elif self.table_titles[title_num] == 'Товары' and not all_articles :
                self.article_col = title_num
                all_articles = self.sheet_reader.col_values(self.article_col)[1:]
            elif self.table_titles[title_num] == 'Группа статуса':
                self.group_status_col = title_num
                
        all_niks = self.sheet_reader.col_values(self.nik_col)[1:]
        print(SUCCESS_MESSAGE+'\t Вытащили оттуда все ники и артикулы')
        return all_niks, all_articles

    def get_freq_dict(self, data:list) -> dict:
        freq_dict = {}
        for item in data:
            if item == '':
                pass
            else:
                count = 0
                for item2 in data:
                    if item == item2:
                        count += 1
                freq_dict[item] = count
        print(SUCCESS_MESSAGE+'\t Создали частнотный словарь')
        return freq_dict
    
    def find_niks_in_sheet(self, data):

        self.nik_worksheet = self.worksheet
        names_col = self.nik_worksheet.find('наименование товара').col
        self.order_names = self.nik_worksheet.col_values(names_col)
        for item in data:  
            print(WARNING_MESSAGE + f'\t Пробуем записать "{item}" в таблицу')
            self.try_write(item, data[item]) 
            # break
        if self.NONE_EXIST_NIKS:
            self.create_non_existent_FILE(self.NONE_EXIST_NIKS, filename='NONE NIKS')
            self.create_non_existent_TABLE(non_existent_file='NONE NIKS',  col=self.nik_col)


    def try_write(self, item, value):
        try:
            count = len(self.order_names)
            for name_row in range(len(self.order_names)):
                if item.strip() in [i.strip() for i in self.order_names[name_row].split('\n')]:
                    
                    total_number_of_orders_value = self.nik_worksheet.acell(f'E{name_row+1}').value if self.nik_worksheet.acell(f'E{name_row+1}').value != None else 0
                    total_number_of_orders_IN_PROCCESING_value = self.nik_worksheet.acell(f'F{name_row+1}').value if self.nik_worksheet.acell(f'F{name_row+1}').value != None else 0
                    total_number_of_orders_SPAM_value = self.nik_worksheet.acell(f'G{name_row+1}').value if self.nik_worksheet.acell(f'G{name_row+1}').value != None else 0
                    total_number_of_orders_CANCEL_value = self.nik_worksheet.acell(f'H{name_row+1}').value if self.nik_worksheet.acell(f'H{name_row+1}').value != None else 0
                    total_number_of_orders_SENT_value = self.nik_worksheet.acell(f'I{name_row+1}').value if self.nik_worksheet.acell(f'I{name_row+1}').value != None else 0
                    total_number_of_orders_IN_WAY_value = self.nik_worksheet.acell(f'J{name_row+1}').value if self.nik_worksheet.acell(f'J{name_row+1}').value != None else 0
                    total_number_of_orders_BOUGHT_OUT_value = self.nik_worksheet.acell(f'K{name_row+1}').value if self.nik_worksheet.acell(f'K{name_row+1}').value != None else 0
                    total_number_of_orders_REFUND_value = self.nik_worksheet.acell(f'L{name_row+1}').value if self.nik_worksheet.acell(f'L{name_row+1}').value != None else 0

                    spam_group, cancel_group, sent_group, processing_group, group_in_way, group_bought_out, group_refund = self.collect_status_orders(item)

                    self.nik_worksheet.update(f'E{name_row+1}', int(total_number_of_orders_value) + int(value))
                    self.nik_worksheet.update(f'F{name_row+1}', int(total_number_of_orders_IN_PROCCESING_value) + int(processing_group))
                    self.nik_worksheet.update(f'G{name_row+1}', int(total_number_of_orders_SPAM_value) + int(spam_group))
                    self.nik_worksheet.update(f'H{name_row+1}', int(total_number_of_orders_CANCEL_value) + int(cancel_group))
                    self.nik_worksheet.update(f'I{name_row+1}', int(total_number_of_orders_SENT_value) + int(sent_group))
                    self.nik_worksheet.update(f'J{name_row+1}', int(total_number_of_orders_IN_WAY_value) + int(group_in_way))
                    self.nik_worksheet.update(f'K{name_row+1}', int(total_number_of_orders_BOUGHT_OUT_value) + int(group_bought_out))
                    self.nik_worksheet.update(f'L{name_row+1}', int(total_number_of_orders_REFUND_value) + int(group_refund))
                    print(SUCCESS_MESSAGE + '\tНайден товар ', item.lower())
                    count -= 1
                    return
            self.NONE_EXIST_NIKS.append(item)
            print(ERROR_MESSAGE+'Не найден ', item)
        # except BaseException:
        #     print(item)

        except gspread.exceptions.APIError:
            time.sleep(31)
            print(ERROR_MESSAGE + '\tБоту нужно отдохнуть 30 секунд')
            self.try_write(item, value)
    
    def collect_status_orders(self, item):            
        spam_group = 0
        cancel_group = 0
        sent_group = 0
        processing_group = 0

        group_in_way = 0
        group_bought_out = 0
        group_refund = 0

        for row_num in range(self.sheet_reader.nrows):
            if item.strip() in self.sheet_reader.row_values(row_num):
                status = self.sheet_reader.cell(row_num, self.group_status_col).value
                if status  == 'Ошибка/Спам/Дубль':
                    spam_group += 1

                elif status == 'Отменен':
                    cancel_group += 1

                elif status == 'Обработка' :
                    processing_group += 1

                elif status in ('Оплачен', 'Отправлен', 'Принят', 'Возврат'):
                    sent_group += 1
                    if status in ('Отправлен', 'Принят'):
                        group_in_way += 1
                    elif status == 'Оплачен':
                        group_bought_out += 1
                    elif status == 'Возврат':
                        group_refund += 1
        print(SUCCESS_MESSAGE+f'\t Собрали всю информацию по статусам ника: {item} ')
        return spam_group, cancel_group, sent_group, processing_group, group_in_way, group_bought_out, group_refund

    def get_articlles_with_statuses(self, data):
        result = {}
        for article in data:
            bought_out = 0
            refund = 0
            for item_num in range(self.sheet_reader.nrows):
                row = self.sheet_reader.row_values(item_num)
                if row[self.article_col] == article and article != '':
                    if row[self.group_status_col] == 'Оплачен':
                        bought_out += 1
                    elif row[self.group_status_col] == 'Возврат':
                        refund += 1
                result[article] = {
                    'Оплачен': bought_out,
                    'Возврат': refund
                }
        return result

    def find_all_articles_in_sheet(self, data):
        arc_col = self.worksheet.find('Артикул').col
        self.google_all_articles = self.worksheet.col_values(arc_col)
        for item in data: #
                if item != ' ':
                    print(WARNING_MESSAGE + '\t Пробуем найти артикул ', item)
                    self.try_write_article(item)
                    # break

    def try_write_article(self, item):
        try:
            for name_row in range(len(self.google_all_articles)):
                if self.google_all_articles[name_row] == item:
                    total_num_of_product_bought_out = self.worksheet.acell(f'Q{name_row+1}').value if self.worksheet.acell(f'Q{name_row+1}').value != None else 0
                    total_num_of_product_refund = self.worksheet.acell(f'R{name_row+1}').value if self.worksheet.acell(f'R{name_row+1}').value != None else 0

                    group_refund, group_bought_out = self.articles_with_status[item]['Возврат'], self.articles_with_status[item]['Оплачен']
                    self.worksheet.update(f'Q{name_row+1}', int(total_num_of_product_bought_out) + int(group_bought_out))
                    self.worksheet.update(f'R{name_row+1}', int(total_num_of_product_refund) + int(group_refund))

                    print(SUCCESS_MESSAGE + '\t Нашли артикул ', item)
                    return
                    
            print(ERROR_MESSAGE + '\t Не удалось найти артикул ', item)
            self.NONE_EXIST_ARTICLES.append(item)

        except gspread.exceptions.APIError:
            print(ERROR_MESSAGE+'\t Бот умер')
            time.sleep(31)
            self.try_write_article(item)

    def collect_articles_status(self, item):
        for title_num in range(len(self.table_titles)):
            if self.table_titles[title_num] == 'Группа статуса':
                articles_status_col = title_num

        group_refund = 0
        group_bought_out = 0
        for row_num in range(self.sheet_reader.nrows):
            if item.strip() in self.sheet_reader.row_values(row_num):
                if self.sheet_reader.cell(row_num, articles_status_col).value == 'Оплачен':
                    group_bought_out += 1
                elif self.sheet_reader.cell(row_num, articles_status_col).value == 'Возврат':
                    group_refund += 1
        return group_refund, group_bought_out

    def create_non_existent_FILE(self, data,  filename):
            file = open(f'Result/{filename}.txt', 'w')
            for item in data:
                file.write(item + '\n')
            file.close()

    def create_non_existent_TABLE(self, non_existent_file, col):
        file = open("Result/"+non_existent_file+'.txt')
        
        res_data = [self.table_titles, ]
        all_products_in_table = self.sheet_reader.col_values(col)
        for item in range(len(all_products_in_table)):
            if all_products_in_table[item] == '':
                res_data.append(self.sheet_reader.row_values(item))
        for product in file:
            for item in range(len(all_products_in_table)):
                if all_products_in_table[item].strip() == product.strip():

                    # if all_status_in_table[item].strip() in ('Возврат', 'Оплачен'):
                    res_data.append(self.sheet_reader.row_values(item))
        
        self.save_table(res_data, non_existent_file)

    def save_table(self, data, filename):
        workbook_writer = xlsxwriter.Workbook("Result/"+filename+'.xlsx')
        worksheet_writer = workbook_writer.add_worksheet()

        for line in range(len(data)):
            for col in range(len(data[line])):
                worksheet_writer.write(line, col, data[line][col])

        workbook_writer.close() 

if __name__ == '__main__':
    bot = CPA('13N8mWPuPGym1WR0jCdXU_GL4AdMZWYdJQZ22xqW9IDw')
    bot.run()

